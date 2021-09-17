import itertools as it

import numpy as np
import openpyxl

# Excel data output
def output_network_excel_sheet(network, wb, ws=None, title=None):
  if ws is None:
    ws = wb.create_sheet(title = title) 
  elif title is not None:
    ws.title = title

  fluorophore_names = [n.name for n in network.nodes]
  num_fluorophores = len(fluorophore_names)

  K_fret = network.get_K_fret()
  k_out = network.get_k_out()
  k_decay = network.get_k_decay()
  
  ws.append([''] + fluorophore_names)
  for j,f in enumerate(fluorophore_names):
    ws.append([f] + [K_fret[i,j] for i in range(num_fluorophores)])
  ws.append(['k_out'] + list(k_out))
  ws.append(['k_decay'] + list(k_decay))

def output_K_fret_excel_sheet(K, fluorophore_names, wb, ws=None, title=None):
  if ws is None:
    ws = wb.create_sheet(title = title) 
  elif title is not None:
    ws.title = title

  num_fluorophores = len(fluorophore_names)

  ws.append([''] + fluorophore_names)
  for i,f in enumerate(fluorophore_names):
    ws.append([f] + list(K[i,:]))

def output_positions_excel_sheet(positions, fluorophore_names, wb, ws=None, title=None):
  if ws is None:
    ws = wb.create_sheet(title = title) 
  elif title is not None:
    ws.title = title

  for f, pos in zip(fluorophore_names, positions):
    ws.append([f] + list(pos))

def output_distances_excel_sheet(distances, fluorophore_names, wb, ws=None, title=None):
  if ws is None:
    ws = wb.create_sheet(title = title) 
  elif title is not None:
    ws.title = title

  ws.append([''] + fluorophore_names)
  for i1, f1 in enumerate(fluorophore_names):
    ws.append([f1] + list(distances[i1,:]))

def output_network_excel(path, fluorophore_names, K, positions = None, network = None, ideal_network = None):
  num_fluorophores = len(fluorophore_names)

  wb = openpyxl.Workbook()

  ws = wb.active

  if ideal_network is not None:
    output_network_excel_sheet(ideal_network, wb, ws=ws, title='Network rates (ideal)')
    ws=None

  if network is not None:
    output_network_excel_sheet(network, wb, ws=ws, title='Network rates')
    ws=None

  output_K_fret_excel_sheet(K, fluorophore_names, wb, ws=ws, title='Rate matrix')
  ws=None

  if positions is not None:
    distances = np.empty((num_fluorophores, num_fluorophores))
    distances[tuple(zip(*it.product(range(num_fluorophores), repeat=2)))] = [np.linalg.norm(positions[i]-positions[j]) for i,j in it.product(range(num_fluorophores), repeat=2)]
    output_positions_excel_sheet(positions, fluorophore_names, wb, ws=ws, title='Positions')
    output_distances_excel_sheet(distances, fluorophore_names, wb, ws=ws, title='Distances')

  wb.save(filename = path)


# MOL2 output
def output_network_mol2(path, fluorophore_names, positions, fluorophore_types, network_name, comments = None):
  num_fluorophores = len(fluorophore_names)

  # convenience type conversion to get better mol2 display colors in chimerax
  fluor_types_mol2 = list(map(lambda t: {'I': 'Cl', 'C': 'Cd', 'O': 'O', 'Q': 'Ag'}.get(t,t), fluorophore_types))

  with open(path, 'w') as outfile:
    if comments is not None:
      [print(comment, file=outfile) for comment in comments]
    print(file=outfile)
  
    print(f'@<TRIPOS>MOLECULE', file=outfile)
    print(f'{network_name}', file=outfile)
    print(f'{num_fluorophores} 0 1 0 0', file=outfile)
    print(f'SMALL', file=outfile)
    print(f'NO_CHARGES', file=outfile)
    print(file=outfile)
    
    print(f'@<TRIPOS>ATOM', file=outfile)
    for i, (f,pos) in enumerate(zip(fluorophore_names, positions)):
      print(f'{i} {f} {pos[0]} {pos[1]} {pos[2]} {fluor_types_mol2[i]} 0 FRETNET', file=outfile)
    print(file=outfile)
  
    print(f'@<TRIPOS>SUBSTRUCTURE', file=outfile)
    print(f'0 FRETNET 0', file=outfile)
